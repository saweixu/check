import io
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import coordinate_to_tuple


# =========================
# CONFIG STREAMLIT
# =========================
st.set_page_config(
    page_title="Athina Logistics Tool",
    page_icon="logo.png",
    layout="wide"
)

if Path("logo.png").exists():
    st.sidebar.image("logo.png", width=200)

st.sidebar.markdown("### Athina Logistics")
st.sidebar.caption("Global Access")


# =========================
# CONSTANTS
# =========================
TEMPLATE_FILE = "T1_SGS.xlsx"

COUNTRY_CODE_RE = re.compile(r"^[A-Z]{2}$")
_REF_RE = re.compile(
    r"^\s*=\s*(?:(?P<sheet>'[^']+'|[A-Za-z0-9 _.-]+)!)?\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)\s*$"
)


# =========================
# HELPERS FINAL CHECK
# =========================
def to_decimal(val):
    if val is None:
        return None
    s = str(val).replace(",", ".").strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def to_decimal_zero(val):
    if val is None:
        return Decimal("0")
    try:
        s = str(val).replace(",", ".").strip()
        if s == "":
            return Decimal("0")
        return Decimal(s)
    except Exception:
        return Decimal("0")


def sheet_by_name_ci(wb, wanted):
    norm = wanted.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == norm:
            return wb[name]
    return None


def find_sum_row(ws, start_row=1, label_col="B"):
    if ws is None:
        return None
    for r in range(start_row, 10000):
        v = ws[f"{label_col}{r}"].value
        if not v:
            continue
        s = str(v).strip().upper()
        if s == "SUM" or re.match(r"^SUM\b", s):
            return r
    return None


def is_cell_in_merged(ws, row, col):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return True
    return False


def get_merged_value(ws, cell_ref):
    cell = ws[cell_ref]
    if cell.value is not None:
        return cell.value

    r, c = coordinate_to_tuple(cell_ref)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return None


def get_effective_cell_value(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return val

    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return None


def contains_chinese(text):
    if not isinstance(text, str):
        return False
    return re.search(r"[\u4e00-\u9fff]", text) is not None


def _col_to_idx(col):
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def resolve_simple_formula(wbX, ws_current, formula, depth=0):
    if depth > 10:
        return None
    if not isinstance(formula, str) or not formula.strip().startswith("="):
        return None

    m = _REF_RE.match(formula)
    if not m:
        return None

    sheet = m.group("sheet")
    col = m.group("col")
    row = int(m.group("row"))

    if sheet:
        sheet = sheet.strip()
        if sheet.startswith("'") and sheet.endswith("'"):
            sheet = sheet[1:-1]
        if sheet not in wbX.sheetnames:
            return None
        ws = wbX[sheet]
    else:
        ws = ws_current

    c = _col_to_idx(col)
    v = ws.cell(row=row, column=c).value
    if isinstance(v, str) and v.strip().startswith("="):
        return resolve_simple_formula(wbX, ws, v, depth + 1)
    return v


def header_value(wb_formula, ws_dataonly, ws_formula, ref):
    v = ws_dataonly[ref].value
    if v is not None and str(v).strip() != "":
        return v

    if ws_formula is None:
        return v

    vf = ws_formula[ref].value
    if isinstance(vf, str) and vf.strip().startswith("="):
        return resolve_simple_formula(wb_formula, ws_formula, vf)

    return vf


def q2(x):
    if x is None:
        return None
    try:
        return x.quantize(Decimal("0.01"))
    except Exception:
        return None


def check_file(uploaded_file):
    errors = []
    warnings = []
    info = {}

    raw = uploaded_file.getvalue()
    bio1 = io.BytesIO(raw)
    bio2 = io.BytesIO(raw)

    wb = load_workbook(bio1, data_only=True)
    wb_f = load_workbook(bio2, data_only=False)

    ws_inv = sheet_by_name_ci(wb, "INVOICE")
    ws_pack = sheet_by_name_ci(wb, "PACKING LIST")
    ws_inv_f = sheet_by_name_ci(wb_f, "INVOICE")
    ws_pack_f = sheet_by_name_ci(wb_f, "PACKING LIST")

    fname = uploaded_file.name
    fname_no_ext = fname.rsplit(".", 1)[0]

    if ws_inv is None:
        errors.append("Missing sheet: INVOICE")
        return build_result(fname, errors, warnings, info)

    if ws_pack is None:
        warnings.append("Missing sheet: PACKING LIST (cross-checks limited)")

    def cell(ws, ref):
        return ws[ref].value if ws is not None else None

    inv_a2 = header_value(wb_f, ws_inv, ws_inv_f, "A2")
    inv_c4 = header_value(wb_f, ws_inv, ws_inv_f, "C4")

    if inv_a2 != inv_c4:
        errors.append(f"INVOICE header mismatch: A2='{inv_a2}' vs C4='{inv_c4}'")

    if ws_pack is not None:
        pack_a2 = header_value(wb_f, ws_pack, ws_pack_f, "A2")
        if not (inv_a2 == inv_c4 == pack_a2):
            errors.append(
                f"Header mismatch across sheets: INVOICE A2='{inv_a2}', C4='{inv_c4}', PACKING A2='{pack_a2}'"
            )

    inv_c5_val = header_value(wb_f, ws_inv, ws_inv_f, "C5")
    inv_j4_val = header_value(wb_f, ws_inv, ws_inv_f, "J4")
    pack_b4_val = header_value(wb_f, ws_pack, ws_pack_f, "B4") if ws_pack is not None else None

    inv_c5 = str(inv_c5_val).strip() if inv_c5_val else ""
    inv_j4 = str(inv_j4_val).strip() if inv_j4_val else ""
    pack_b4 = str(pack_b4_val).strip() if pack_b4_val else ""

    if ws_pack is not None:
        if not (fname_no_ext == inv_c5 == inv_j4 == pack_b4):
            errors.append(
                f"Filename mismatch: file='{fname_no_ext}', INVOICE C5='{inv_c5}', INVOICE J4='{inv_j4}', PACKING B4='{pack_b4}'"
            )
    else:
        if not (fname_no_ext == inv_c5 == inv_j4):
            errors.append(
                f"Filename mismatch: file='{fname_no_ext}', INVOICE C5='{inv_c5}', INVOICE J4='{inv_j4}'"
            )

    j13 = str(cell(ws_inv, "J13")).strip().upper() if cell(ws_inv, "J13") else ""
    if j13 != "EUR":
        errors.append(f"J13 must be 'EUR' (found: {cell(ws_inv, 'J13')})")

    j14 = str(cell(ws_inv, "J14")).strip().upper() if cell(ws_inv, "J14") else ""
    if j14 != "CIF":
        warnings.append(f"J14 should be 'CIF' (found: {cell(ws_inv, 'J14')})")

    try:
        j16 = Decimal(str(cell(ws_inv, "J16")).strip()) if cell(ws_inv, "J16") else None
    except InvalidOperation:
        j16 = None

    if j16 != Decimal(4200):
        errors.append(f"J16 must be 4200 (found: {cell(ws_inv, 'J16')})")

    for r in range(11, 16):
        v = cell(ws_inv, f"C{r}")
        if v is None or str(v).strip() == "":
            errors.append(f"C{r} is empty")

    for r in (16, 17):
        v = cell(ws_inv, f"C{r}")
        s = str(v).strip().upper() if v is not None else ""
        if not COUNTRY_CODE_RE.match(s):
            errors.append(f"C{r} must be a 2-letter country code (found: '{v}')")

    for r in (11, 13, 15):
        v = cell(ws_inv, f"J{r}")
        if v is None or str(v).strip() == "":
            errors.append(f"J{r} is empty")

    c9 = str(cell(ws_inv, "C9")).strip().upper() if cell(ws_inv, "C9") else ""
    if c9 != "CN":
        errors.append(f"C9 must be 'CN' (found: {cell(ws_inv, 'C9')})")

    inv_sum_row = find_sum_row(ws_inv, start_row=19, label_col="B")
    pack_sum_row = find_sum_row(ws_pack, start_row=6, label_col="B") if ws_pack else None

    if inv_sum_row is None:
        errors.append("INVOICE SUM row not found in column B")

    if ws_pack is not None and pack_sum_row is None:
        errors.append("PACKING LIST SUM row not found in column B")

    if inv_sum_row:
        inv_ch_rows = []
        for row in range(20, inv_sum_row):
            if contains_chinese(ws_inv[f"B{row}"].value):
                inv_ch_rows.append(f"B{row}")

        if inv_ch_rows:
            errors.append(
                f"Chinese characters found in INVOICE descriptions: {', '.join(inv_ch_rows[:20])}"
                + (" ..." if len(inv_ch_rows) > 20 else "")
            )

    if ws_pack and pack_sum_row:
        pack_ch_rows = []
        for row in range(6, pack_sum_row):
            if contains_chinese(ws_pack[f"B{row}"].value):
                pack_ch_rows.append(f"B{row}")

        if pack_ch_rows:
            errors.append(
                f"Chinese characters found in PACKING LIST descriptions: {', '.join(pack_ch_rows[:20])}"
                + (" ..." if len(pack_ch_rows) > 20 else "")
            )

    if inv_sum_row:
        merged_inv = []
        for r in range(20, inv_sum_row):
            if is_cell_in_merged(ws_inv, r, 10):
                merged_inv.append(f"J{r}")
            if is_cell_in_merged(ws_inv, r, 11):
                merged_inv.append(f"K{r}")

        if merged_inv:
            errors.append(
                "Merged cells not allowed in INVOICE J/K area: "
                + ", ".join(merged_inv[:20])
                + (" ..." if len(merged_inv) > 20 else "")
            )

    if ws_pack and pack_sum_row:
        merged_pack = []
        for r in range(6, pack_sum_row):
            if is_cell_in_merged(ws_pack, r, 9):
                merged_pack.append(f"I{r}")
            if is_cell_in_merged(ws_pack, r, 10):
                merged_pack.append(f"J{r}")

        if merged_pack:
            warnings.append(
                "Merged cells found in PACKING LIST I/J area. SGS generator will handle merged net/gross cells: "
                + ", ".join(merged_pack[:20])
                + (" ..." if len(merged_pack) > 20 else "")
            )

    if inv_sum_row:
        for r in range(20, inv_sum_row):
            val = ws_inv[f"G{r}"].value
            if isinstance(val, str) and len(val.strip()) > 48:
                errors.append("INVOICE column G contains a value longer than 48 characters")
                break

    if inv_sum_row:
        di_errors = []

        def check_DI_cell(row, col_letter, col_index):
            if is_cell_in_merged(ws_inv, row, col_index):
                return

            val = ws_inv[f"{col_letter}{row}"].value

            if val is None:
                di_errors.append(f"{col_letter}{row} empty")
                return

            s = str(val).strip().replace(",", ".")

            if s in ("0", "0.0", "0.00"):
                di_errors.append(f"{col_letter}{row} = 0")
                return

            try:
                Decimal(s)
            except InvalidOperation:
                di_errors.append(f"{col_letter}{row} non-numeric ('{val}')")

        for r in range(20, inv_sum_row):
            check_DI_cell(r, "D", 4)
            check_DI_cell(r, "I", 9)

        if di_errors:
            errors.append(
                "Invalid goods values in INVOICE D/I: "
                + "; ".join(di_errors[:20])
                + (" ..." if len(di_errors) > 20 else "")
            )

    if inv_sum_row:
        text_rows = []
        bad_rows = []

        for r in range(20, inv_sum_row):
            j_val = ws_inv[f"J{r}"].value
            k_val = ws_inv[f"K{r}"].value

            if isinstance(j_val, str) or isinstance(k_val, str):
                text_rows.append(r)
                continue

            j_dec = to_decimal(j_val)
            k_dec = to_decimal(k_val)

            if j_dec is not None and k_dec is not None and j_dec >= k_dec:
                bad_rows.append(r)

        if text_rows:
            errors.append(
                f"Text found in INVOICE J/K line weights: rows {text_rows[:20]}"
                + (" ..." if len(text_rows) > 20 else "")
            )

        if bad_rows:
            errors.append(
                f"Net weight >= gross weight in INVOICE: rows {bad_rows[:20]}"
                + (" ..." if len(bad_rows) > 20 else "")
            )

    if inv_sum_row and ws_pack and pack_sum_row:
        inv_pieces = to_decimal(ws_inv[f"H{inv_sum_row}"].value)
        inv_net = to_decimal(ws_inv[f"J{inv_sum_row}"].value)
        inv_gross = to_decimal(ws_inv[f"K{inv_sum_row}"].value)

        pack_pieces = to_decimal(ws_pack[f"H{pack_sum_row}"].value)
        pack_net = to_decimal(ws_pack[f"I{pack_sum_row}"].value)
        pack_gross = to_decimal(ws_pack[f"J{pack_sum_row}"].value)
        pack_cartons = to_decimal(get_merged_value(ws_pack, f"G{pack_sum_row}"))

        if inv_net is None or inv_gross is None:
            errors.append("INVOICE total net/gross not numeric at SUM row")

        if pack_net is None or pack_gross is None:
            errors.append("PACKING LIST total net/gross not numeric at SUM row")

        if inv_pieces != pack_pieces:
            errors.append(f"Total pieces mismatch: INVOICE={inv_pieces}, PACKING={pack_pieces}")

        if inv_net != pack_net:
            errors.append(f"Total net weight mismatch: INVOICE={inv_net}, PACKING={pack_net}")

        if inv_gross != pack_gross:
            errors.append(f"Total gross weight mismatch: INVOICE={inv_gross}, PACKING={pack_gross}")

        if inv_net is not None and inv_gross is not None and inv_net > inv_gross:
            errors.append(f"INVOICE total net weight > gross weight ({inv_net} > {inv_gross})")

        if pack_cartons is None:
            errors.append("PACKING LIST total cartons missing or non-numeric")
        else:
            info["cartons"] = str(pack_cartons)

        if pack_gross is not None:
            info["gross_weight"] = str(pack_gross)

        inv_b_values = [
            str(ws_inv[f"B{r}"].value).strip() if ws_inv[f"B{r}"].value else ""
            for r in range(20, inv_sum_row)
        ]
        pack_b_values = [
            str(ws_pack[f"B{r}"].value).strip() if ws_pack[f"B{r}"].value else ""
            for r in range(6, pack_sum_row)
        ]

        if inv_b_values != pack_b_values:
            errors.append(
                f"Description column B mismatch between INVOICE and PACKING LIST "
                f"(INV lines={len(inv_b_values)}, PACK lines={len(pack_b_values)})"
            )

        line_errors = []
        offset = 14

        for inv_row in range(20, inv_sum_row):
            pack_row = inv_row - offset

            if pack_row < 6 or pack_row >= pack_sum_row:
                continue

            inv_pieces_line = to_decimal(ws_inv[f"H{inv_row}"].value)
            inv_net_line = q2(to_decimal(ws_inv[f"J{inv_row}"].value))
            inv_gross_line = q2(to_decimal(ws_inv[f"K{inv_row}"].value))

            pack_pieces_line = to_decimal(ws_pack[f"H{pack_row}"].value)
            pack_net_line = q2(to_decimal(ws_pack[f"I{pack_row}"].value))
            pack_gross_line = q2(to_decimal(ws_pack[f"J{pack_row}"].value))

            if inv_pieces_line != pack_pieces_line:
                line_errors.append(f"Row {inv_row}: pieces mismatch")

            if inv_net_line != pack_net_line:
                line_errors.append(f"Row {inv_row}: net weight mismatch")

            if inv_gross_line != pack_gross_line:
                line_errors.append(f"Row {inv_row}: gross weight mismatch")

        if line_errors:
            errors.append(
                "Line-by-line differences between INVOICE and PACKING LIST: "
                + "; ".join(line_errors[:25])
                + (" ..." if len(line_errors) > 25 else "")
            )

    if ws_pack and pack_sum_row:
        bad_rows = []

        for r in range(6, pack_sum_row):
            gv_raw = get_effective_cell_value(ws_pack, r, 7)
            gv_dec = to_decimal(gv_raw)

            if gv_raw is None or str(gv_raw).strip() == "":
                bad_rows.append(f"G{r}=empty")
                continue

            if gv_dec is None:
                bad_rows.append(f"G{r}=non-numeric")
                continue

            if gv_dec == 0:
                bad_rows.append(f"G{r}=0")

        if bad_rows:
            errors.append(
                "Invalid carton values in PACKING LIST column G: "
                + "; ".join(bad_rows[:25])
                + (" ..." if len(bad_rows) > 25 else "")
            )

    return build_result(fname, errors, warnings, info)


def build_result(fname, errors, warnings, info):
    status = "OK"
    if errors:
        status = "ERROR"
    elif warnings:
        status = "WARNING"

    return {
        "file": fname,
        "status": status,
        "errors": errors,
        "warnings": warnings,
        "info": info,
        "error_count": len(errors),
        "warning_count": len(warnings),
    }


# =========================
# SGS GENERATION - SMART VERSION
# Logic replaced with summary_hs_code.py aggregation,
# but output remains ONLY the SGS template file.
# =========================
def sgs_to_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    s = s.replace(" ", "")
    if s.count(",") and s.count("."):
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def sgs_find_header_row_by_keyword(ws, col, keyword):
    needle = str(keyword).strip().lower()
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, str) and needle in v.strip().lower():
            return row
    return None


def sgs_merged_top_left(ws, row, col):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col, str(rng)
    return row, col, None


def sgs_get_cell_value_merged(ws, row, col):
    r0, c0, rng = sgs_merged_top_left(ws, row, col)
    return ws.cell(row=r0, column=c0).value, rng


def sgs_norm_desc(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = " ".join(s.split())
    s = "".join(ch if ch.isalnum() or ch.isspace() else " " for ch in s)
    return " ".join(s.split())


def sgs_parse_invoice(ws):
    sub_order_no = ws["C5"].value or ""
    header_row = sgs_find_header_row_by_keyword(ws, col=2, keyword="Description of Goods")
    if header_row is None:
        return [], str(sub_order_no).strip()

    products = []
    for r in range(header_row + 1, ws.max_row + 1):
        desc = ws.cell(row=r, column=2).value
        if desc is None:
            break
        if isinstance(desc, str) and desc.strip().upper() == "SUM":
            break

        hs = ws.cell(row=r, column=3).value
        val = ws.cell(row=r, column=9).value
        mark = ws.cell(row=r, column=7).value

        if hs is None or desc is None:
            continue

        products.append({
            "hs_code": str(hs).strip(),
            "desc": str(desc).strip(),
            "custom_value": sgs_to_float(val),
            "mark": str(mark).strip() if mark not in (None, "") else "",
            "sub_order_no": str(sub_order_no).strip(),
        })

    return products, str(sub_order_no).strip()


def sgs_build_invoice_index(products):
    from collections import defaultdict
    by_mark_desc_seq = defaultdict(list)
    by_desc = defaultdict(list)

    for p in products:
        if p.get("mark"):
            by_mark_desc_seq[(p["mark"], p["desc"])].append(p["hs_code"])
        by_desc[p["desc"]].append(p["hs_code"])

    return by_mark_desc_seq, by_desc


def sgs_parse_packing_list_rows(ws):
    header_row = sgs_find_header_row_by_keyword(ws, col=2, keyword="Description of Goods")
    if header_row is None:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        desc = ws.cell(row=r, column=2).value
        if desc is None:
            break
        if isinstance(desc, str) and desc.strip().upper() == "SUM":
            break

        mark_val, _ = sgs_get_cell_value_merged(ws, r, 5)  # E
        carton_val, carton_rng = sgs_get_cell_value_merged(ws, r, 7)  # G
        net_val, net_rng = sgs_get_cell_value_merged(ws, r, 9)  # I
        gross_val, gross_rng = sgs_get_cell_value_merged(ws, r, 10)  # J

        rows.append({
            "desc": str(desc).strip(),
            "mark": str(mark_val).strip() if mark_val not in (None, "") else "",
            "carton": sgs_to_float(carton_val),
            "net": sgs_to_float(net_val),
            "gross": sgs_to_float(gross_val),
            "carton_rng": carton_rng or f"R{r}C7",
            "net_rng": net_rng or f"R{r}C9",
            "gross_rng": gross_rng or f"R{r}C10",
        })

    return rows


def sgs_collect_from_uploaded_file(uploaded_file):
    from collections import Counter

    raw = uploaded_file.getvalue()
    wb = load_workbook(io.BytesIO(raw), data_only=True)

    ws_inv = sheet_by_name_ci(wb, "INVOICE")
    ws_pl = sheet_by_name_ci(wb, "PACKING LIST")

    if ws_inv is None or ws_pl is None:
        return [], [], []

    inv_products, sub_order_no = sgs_parse_invoice(ws_inv)
    by_mark_desc_seq, by_desc = sgs_build_invoice_index(inv_products)
    pl_rows = sgs_parse_packing_list_rows(ws_pl)

    groups = []
    idx_map = {}
    for pr in pl_rows:
        crng = pr["carton_rng"]
        if crng not in idx_map:
            idx_map[crng] = len(groups)
            groups.append({"carton_rng": crng, "rows": []})
        groups[idx_map[crng]]["rows"].append(pr)

    group_calc = []
    pl_md_counter = Counter()

    for g in groups:
        rows = g["rows"]
        if not rows:
            continue

        total_cartons = float(rows[0]["carton"] or 0.0)
        n = len(rows)

        hs_list = []
        for pr in rows:
            if pr["mark"] and (pr["mark"], pr["desc"]) in by_mark_desc_seq:
                seq = by_mark_desc_seq[(pr["mark"], pr["desc"])]
                idx = pl_md_counter[(pr["mark"], pr["desc"])]
                if idx >= len(seq):
                    idx = len(seq) - 1
                hs = seq[idx] if seq else "PL_UNMATCHED"
                pl_md_counter[(pr["mark"], pr["desc"])] += 1
            else:
                cands = by_desc.get(pr["desc"], [])
                if cands:
                    freq = {}
                    for h in cands:
                        freq[h] = freq.get(h, 0) + 1
                    hs = sorted(freq.items(), key=lambda x: (-x[1], x[0]))[0][0]
                else:
                    hs = "PL_UNMATCHED"
            hs_list.append(hs)

        cartons_assigned = [0.0] * n
        if total_cartons <= 0:
            pass
        elif total_cartons >= n:
            for i in range(n - 1):
                cartons_assigned[i] = 1.0
            cartons_assigned[n - 1] = float(total_cartons - (n - 1))
        else:
            for i in range(int(total_cartons)):
                cartons_assigned[i] = 1.0

        recipient_idx = None
        for i in range(n - 1, -1, -1):
            if cartons_assigned[i] > 0:
                recipient_idx = i
                break

        group_calc.append({
            "rows": rows,
            "hs_list": hs_list,
            "cartons_assigned": cartons_assigned,
            "recipient_idx": recipient_idx,
            "total_cartons": total_cartons,
        })

    # SMART HOLDER v2 from summary_hs_code.py
    group_appear = []
    for gc in group_calc:
        seen = set()
        for i, pr in enumerate(gc["rows"]):
            seen.add((gc["hs_list"][i], sgs_norm_desc(pr["desc"])))
        group_appear.append(seen)

    occ_counter = Counter()
    for seen in group_appear:
        for k in seen:
            occ_counter[k] += 1

    assigned_carton_so_far = Counter()

    for gc in group_calc:
        rows = gc["rows"]
        hs_list = gc["hs_list"]
        n = len(rows)
        total_cartons = int(gc["total_cartons"] or 0)
        cartons_assigned = [0.0] * n

        if total_cartons <= 0:
            pass
        elif total_cartons >= n:
            for i in range(n):
                cartons_assigned[i] = 1.0
            extra = total_cartons - n
            if extra > 0:
                cartons_assigned[0] += float(extra)
        else:
            weights = []
            for i, pr in enumerate(rows):
                k = (hs_list[i], sgs_norm_desc(pr["desc"]))
                net_hint = float(pr.get("net") or 0.0)
                weights.append((assigned_carton_so_far[k], occ_counter[k], -net_hint, i))
            weights.sort()
            for idx in [w[-1] for w in weights[:total_cartons]]:
                cartons_assigned[idx] = 1.0

        recipient_idx = None
        for i in range(n):
            if cartons_assigned[i] > 0:
                recipient_idx = i
                break

        for i in range(n):
            if cartons_assigned[i] > 0:
                k = (hs_list[i], sgs_norm_desc(rows[i]["desc"]))
                assigned_carton_so_far[k] += cartons_assigned[i]

        gc["cartons_assigned"] = cartons_assigned
        gc["recipient_idx"] = recipient_idx

    used_gross_rng = set()
    used_net_rng = set()
    pl_lines = []
    transfers = []

    for gc in group_calc:
        rows = gc["rows"]
        hs_list = gc["hs_list"]
        cartons_assigned = gc["cartons_assigned"]
        recipient_idx = gc.get("recipient_idx")

        if recipient_idx is not None:
            r_key = (hs_list[recipient_idx], rows[recipient_idx]["desc"])
        else:
            r_key = None

        for i, pr in enumerate(rows):
            net_contrib = float(pr.get("net") or 0.0)
            if pr.get("net_rng") in used_net_rng:
                net_contrib = 0.0
            else:
                used_net_rng.add(pr.get("net_rng"))

            gross_contrib = float(pr.get("gross") or 0.0)
            if pr.get("gross_rng") in used_gross_rng:
                gross_contrib = 0.0
            else:
                used_gross_rng.add(pr.get("gross_rng"))

            pl_lines.append((hs_list[i], pr["desc"], cartons_assigned[i], net_contrib, gross_contrib, sub_order_no))

            if cartons_assigned[i] == 0.0 and r_key is not None and i != recipient_idx:
                transfers.append(((hs_list[i], pr["desc"]), r_key, gross_contrib))

    value_lines = []
    for p in inv_products:
        value_lines.append((p.get("hs_code", ""), p.get("desc", ""), p.get("custom_value", 0.0), sub_order_no))

    return value_lines, pl_lines, transfers


def sgs_aggregate(value_lines, pl_lines, transfers=None):
    from collections import defaultdict
    agg = defaultdict(lambda: {"carton": 0.0, "net": 0.0, "gross": 0.0, "value": 0.0, "invoices": set()})

    for hs, desc, val, inv in value_lines:
        key = (hs, desc)
        agg[key]["value"] += float(val or 0.0)
        if inv:
            agg[key]["invoices"].add(inv)

    for hs, desc, carton, net, gross, inv in pl_lines:
        key = (hs, desc)
        agg[key]["carton"] += float(carton or 0.0)
        agg[key]["net"] += float(net or 0.0)
        agg[key]["gross"] += float(gross or 0.0)
        if inv:
            agg[key]["invoices"].add(inv)

    if transfers:
        for (from_hs, from_desc), (to_hs, to_desc), amt in transfers:
            try:
                amt = float(amt or 0.0)
            except Exception:
                continue
            if amt == 0:
                continue

            from_key = (from_hs, from_desc)
            to_key = (to_hs, to_desc)

            if from_key in agg and float(agg[from_key]["carton"] or 0.0) == 0.0:
                agg[from_key]["gross"] -= amt
                agg[to_key]["gross"] += amt

        for key, v in agg.items():
            if float(v["carton"] or 0.0) == 0.0:
                v["gross"] = 0.0
            elif v["gross"] < 0:
                v["gross"] = 0.0

    return agg


def _parse_sub_orders(inv_str: str) -> str:
    if not inv_str:
        return ""
    parts = [p.strip() for p in str(inv_str).split(",") if p.strip()]
    nums = []
    for p in parts:
        m = re.search(r"(\d+)\s*$", p)
        if m:
            nums.append(str(int(m.group(1))))
        else:
            m2 = re.search(r"(\d+)", p)
            if m2:
                nums.append(str(int(m2.group(1))))
    seen = set()
    out = []
    for n in nums:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return " ".join(out)


def safe_sort_invoice_numbers(nums):
    def sort_key(x):
        try:
            return (0, int(str(x)))
        except Exception:
            return (1, str(x))
    return sorted(nums, key=sort_key)


def detect_container_prefix(uploaded_files):
    for f in uploaded_files:
        if f.name.lower().endswith((".xlsx", ".xlsm")):
            return Path(f.name).stem.split("-")[0]
    return None


def _copy_template_row_style(ws, target_row, template_row=3, max_col=66):
    from copy import copy
    for c in range(1, max_col + 1):
        src = ws.cell(template_row, c)
        dst = ws.cell(target_row, c)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def _clear_sgs_rows(ws, start_row=3, max_col=66):
    max_row = max(ws.max_row, start_row + 500)
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def create_sgs_file(uploaded_files, terminal_choice, bl_number):
    template_path = Path(TEMPLATE_FILE)

    if not template_path.exists():
        raise FileNotFoundError(f"Template introuvable : {TEMPLATE_FILE}")

    container_prefix = detect_container_prefix(uploaded_files)
    if not container_prefix:
        raise ValueError("Impossible de détecter le numéro de conteneur depuis les fichiers uploadés.")

    all_value_lines = []
    all_pl_lines = []
    all_transfers = []

    for uploaded_file in uploaded_files:
        if not uploaded_file.name.lower().endswith((".xlsx", ".xlsm")):
            continue
        v_lines, pl_lines, transfers = sgs_collect_from_uploaded_file(uploaded_file)
        all_value_lines.extend(v_lines)
        all_pl_lines.extend(pl_lines)
        all_transfers.extend(transfers)

    agg = sgs_aggregate(all_value_lines, all_pl_lines, all_transfers)

    wb_template = load_workbook(template_path)
    ws = wb_template.active

    start_row = 3
    _clear_sgs_rows(ws, start_row=start_row, max_col=66)

    items = []
    for (hs, desc) in sorted(agg.keys(), key=lambda k: (k[0], k[1])):
        d = agg[(hs, desc)]
        if not str(hs or "").strip() or not str(desc or "").strip():
            continue
        items.append({
            "hs": str(hs).strip(),
            "desc": str(desc).strip(),
            "carton": round(float(d.get("carton", 0.0) or 0.0), 2),
            "net": round(float(d.get("net", 0.0) or 0.0), 2),
            "gross": round(float(d.get("gross", 0.0) or 0.0), 2),
            "value": round(float(d.get("value", 0.0) or 0.0), 2),
            "sub_orders": _parse_sub_orders(", ".join(safe_sort_invoice_numbers(d.get("invoices", set()) or []))),
        })

    lines_count = len(items)
    if lines_count == 0:
        raise ValueError("Aucune ligne SGS générée. Vérifie les fichiers INVOICE / PACKING LIST.")

    year = datetime.now(ZoneInfo("Europe/Brussels")).year

    for i, it in enumerate(items):
        r = start_row + i
        _copy_template_row_style(ws, r, template_row=3, max_col=66)

        ws[f"A{r}"] = it["hs"]
        ws[f"B{r}"] = it["desc"]
        ws[f"G{r}"] = it["value"]
        ws[f"I{r}"] = it["net"]
        ws[f"J{r}"] = it["gross"]
        ws[f"AE{r}"] = it["carton"]
        ws[f"AO{r}"] = it["sub_orders"]

        ws[f"H{r}"] = "EUR"
        ws[f"K{r}"] = "SGS0001863312"
        ws[f"M{r}"] = "NL-Dutch"
        ws[f"P{r}"] = "Rotterdam"
        ws[f"R{r}"] = "BE0417688928"
        ws[f"T{r}"] = "BE-Belgian"
        ws[f"U{r}"] = "Hawe"
        ws[f"V{r}"] = "Kruiningenstraat 188"
        ws[f"W{r}"] = "Schoten"
        ws[f"X{r}"] = 2900
        ws[f"AC{r}"] = "NM"
        ws[f"AD{r}"] = "CT-Carton"
        ws[f"AF{r}"] = "N705-Bill of lading"
        ws[f"AG{r}"] = bl_number.strip()
        ws[f"AN{r}"] = "N380-Commercial Invoice"
        ws[f"AR{r}"] = year
        ws[f"AS{r}"] = "N730-Road consignment note"
        ws[f"AT{r}"] = container_prefix
        ws[f"BL{r}"] = "CN-China"
        ws[f"BM{r}"] = "BE-Belgian"
        ws[f"BN{r}"] = container_prefix

        if terminal_choice == "Delta":
            ws[f"N{r}"] = "E.C.T. DELTA B.V."
            ws[f"O{r}"] = "EUROPEAWEG 875"
            ws[f"Q{r}"] = "3199 LD"
        elif terminal_choice == "Euromax":
            ws[f"N{r}"] = "EUROMAX"
            ws[f"O{r}"] = "MAASVLAKTEWEG 951"
            ws[f"Q{r}"] = "3199 LZ"
        elif terminal_choice == "Empty":
            ws[f"N{r}"] = ""
            ws[f"O{r}"] = ""
            ws[f"Q{r}"] = ""

    date_str = datetime.now(ZoneInfo("Europe/Brussels")).strftime("%Y%m%d")
    out_name = f"T1 Request-{container_prefix}-ATH-{date_str}-{lines_count}lines.xlsx"

    output = io.BytesIO()
    wb_template.save(output)
    output.seek(0)

    return out_name, output.getvalue(), lines_count




def build_sgs_preview(uploaded_files):
    """Return aggregated SGS rows for on-screen preview, using the same logic as final output."""
    all_value_lines = []
    all_pl_lines = []
    all_transfers = []

    for uploaded_file in uploaded_files:
        if not uploaded_file.name.lower().endswith((".xlsx", ".xlsm")):
            continue
        v_lines, pl_lines, transfers = sgs_collect_from_uploaded_file(uploaded_file)
        all_value_lines.extend(v_lines)
        all_pl_lines.extend(pl_lines)
        all_transfers.extend(transfers)

    agg = sgs_aggregate(all_value_lines, all_pl_lines, all_transfers)
    rows = []
    for (hs, desc) in sorted(agg.keys(), key=lambda k: (k[0], k[1])):
        d = agg[(hs, desc)]
        if not str(hs or "").strip() or not str(desc or "").strip():
            continue
        rows.append({
            "HS code": str(hs).strip(),
            "Description": str(desc).strip(),
            "Carton": round(float(d.get("carton", 0.0) or 0.0), 2),
            "Net": round(float(d.get("net", 0.0) or 0.0), 2),
            "Gross": round(float(d.get("gross", 0.0) or 0.0), 2),
            "Value": round(float(d.get("value", 0.0) or 0.0), 2),
            "Invoices": _parse_sub_orders(", ".join(safe_sort_invoice_numbers(d.get("invoices", set()) or []))),
        })
    return rows

# =========================
# UI
# =========================
st.title("Final Check + SGS Generator")
st.caption("Upload all invoice Excel files. SGS uses the smart merged-carton logic from summary_hs_code.py and generates only the SGS template file.")

uploaded_files = st.file_uploader(
    "Upload invoice Excel files (.xlsx)",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
)

if uploaded_files:
    results = []

    for f in uploaded_files:
        try:
            results.append(check_file(f))
        except Exception as e:
            results.append({
                "file": f.name,
                "status": "ERROR",
                "errors": [f"File could not be read: {e}"],
                "warnings": [],
                "info": {},
                "error_count": 1,
                "warning_count": 0,
            })

    df = pd.DataFrame([
        {
            "File": r["file"],
            "Status": r["status"],
            "Errors": r["error_count"],
            "Warnings": r["warning_count"],
            "Cartons": r["info"].get("cartons", ""),
            "Gross Weight": r["info"].get("gross_weight", ""),
        }
        for r in results
    ])

    total_files = len(results)
    files_with_errors = sum(1 for r in results if r["error_count"] > 0)
    files_with_warnings_only = sum(
        1 for r in results if r["error_count"] == 0 and r["warning_count"] > 0
    )
    files_ok = sum(
        1 for r in results if r["error_count"] == 0 and r["warning_count"] == 0
    )

    total_cartons = Decimal("0")
    total_gross_weight = Decimal("0")

    for r in results:
        cartons = to_decimal(r["info"].get("cartons", ""))
        gross = to_decimal(r["info"].get("gross_weight", ""))

        if cartons is not None:
            total_cartons += cartons

        if gross is not None:
            total_gross_weight += gross

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Files checked", total_files)
    c2.metric("Files with errors", files_with_errors)
    c3.metric("Files with warnings only", files_with_warnings_only)
    c4.metric("Total cartons", f"{total_cartons:,.0f}")
    c5.metric("Total gross weight", f"{total_gross_weight:,.2f}")

    if files_ok:
        st.success(f"{files_ok} file(s) passed with no issues.")

    if files_with_warnings_only:
        st.warning(f"{files_with_warnings_only} file(s) have warnings only.")

    if files_with_errors:
        st.error(f"{files_with_errors} file(s) have errors. SGS generation blocked.")

    st.subheader("Summary")
    st.dataframe(df, use_container_width=True, hide_index=True)

    problem_files = [
        r for r in results
        if r["error_count"] > 0 or r["warning_count"] > 0
    ]

    st.subheader("Files with issues")

    if not problem_files:
        st.info("No issues found.")
    else:
        for r in problem_files:
            icon = "❌" if r["error_count"] > 0 else "⚠️"
            title = f"{icon} {r['file']}  |  Errors: {r['error_count']}  |  Warnings: {r['warning_count']}"

            with st.expander(title, expanded=False):
                if r["errors"]:
                    st.markdown("**Errors**")
                    for msg in r["errors"]:
                        st.write(f"- {msg}")

                if r["warnings"]:
                    st.markdown("**Warnings**")
                    for msg in r["warnings"]:
                        st.write(f"- {msg}")

    st.divider()

    if files_with_errors > 0:
        st.error("SGS step stopped because at least one invoice has errors.")

    else:
        st.subheader("SGS Generation")

        if not Path(TEMPLATE_FILE).exists():
            st.error(f"Template missing: {TEMPLATE_FILE}")
            st.info("Put T1_SGS.xlsx in the same folder as app.py.")
        else:
            terminal_choice = st.radio(
                "Terminal",
                ["Delta", "Euromax", "Empty"],
                horizontal=True,
            )

            bl_number = st.text_input("BL number")

            with st.expander("Preview SGS calculation", expanded=False):
                try:
                    preview_rows = build_sgs_preview(uploaded_files)
                    if preview_rows:
                        preview_df = pd.DataFrame(preview_rows)
                        st.dataframe(preview_df, use_container_width=True, hide_index=True)
                        st.caption(
                            f"Generated lines: {len(preview_rows)} | "
                            f"Cartons: {preview_df['Carton'].sum():,.2f} | "
                            f"Gross: {preview_df['Gross'].sum():,.2f} | "
                            f"Net: {preview_df['Net'].sum():,.2f} | "
                            f"Value: {preview_df['Value'].sum():,.2f}"
                        )
                    else:
                        st.info("No SGS lines found for preview.")
                except Exception as e:
                    st.warning(f"Preview unavailable: {e}")

            if st.button("Generate SGS file", type="primary"):
                try:
                    out_name, sgs_bytes, lines_count = create_sgs_file(
                        uploaded_files=uploaded_files,
                        terminal_choice=terminal_choice,
                        bl_number=bl_number,
                    )

                    st.success(f"SGS file generated successfully: {lines_count} line(s).")

                    st.download_button(
                        label="Download SGS file",
                        data=sgs_bytes,
                        file_name=out_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                except Exception as e:
                    st.error(f"Could not generate SGS file: {e}")
